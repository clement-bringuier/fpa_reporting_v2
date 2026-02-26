import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
import os

# ─────────────────────────────────────────────
# COULEURS & STYLES
# ─────────────────────────────────────────────

# Lignes de total/marge → fond gris foncé, texte blanc, bold
STYLE_TOTAL = {
    "fill": PatternFill("solid", fgColor="2F3640"),
    "font": Font(bold=True, color="FFFFFF", size=10),
}

# Lignes de marge clé (Gross Margin, EBITDA, EBIT) → fond bleu, texte blanc, bold
STYLE_MARGIN = {
    "fill": PatternFill("solid", fgColor="1A6B9A"),
    "font": Font(bold=True, color="FFFFFF", size=10),
}

# En-tête colonnes → fond gris moyen, texte blanc, bold
STYLE_HEADER = {
    "fill": PatternFill("solid", fgColor="636E72"),
    "font": Font(bold=True, color="FFFFFF", size=10),
}

# Lignes de détail → fond blanc, texte noir
STYLE_DETAIL = {
    "fill": PatternFill("solid", fgColor="FFFFFF"),
    "font": Font(color="000000", size=10),
}

# Lignes de détail alternées → fond gris très clair
STYLE_DETAIL_ALT = {
    "fill": PatternFill("solid", fgColor="F5F6FA"),
    "font": Font(color="000000", size=10),
}

# Marges clés pour style spécial
KEY_MARGINS = {"Gross Margin", "Contribution Margin", "EBITDA", "EBIT"}

# Format nombre
NUMBER_FORMAT = '#,##0;[Red]-#,##0'

thin = Side(style="thin", color="D0D0D0")
BORDER_THIN = Border(bottom=thin)
BORDER_TOTAL = Border(
    top=Side(style="medium", color="000000"),
    bottom=Side(style="medium", color="000000")
)


# ─────────────────────────────────────────────
# ÉCRITURE D'UN ONGLET P&L
# ─────────────────────────────────────────────

def write_pl_sheet(ws, pl_df: pd.DataFrame, sheet_title: str, period_label: str):
    """
    Écrit un P&L dans un onglet Excel avec mise en forme.
    
    Args:
        ws: openpyxl worksheet
        pl_df: DataFrame avec index = labels P&L, colonne "Montant"
        sheet_title: titre affiché dans la cellule A1
        period_label: ex "Décembre 2025"
    """
    from src.pl_builder import PL_STRUCTURE

    # Ligne 1 : titre
    ws["A1"] = sheet_title
    ws["A1"].font = Font(bold=True, size=13, color="1A6B9A")
    ws["A1"].alignment = Alignment(horizontal="left")

    # Ligne 2 : période
    ws["A2"] = period_label
    ws["A2"].font = Font(italic=True, size=10, color="636E72")

    # Ligne 3 : vide

    # Ligne 4 : en-tête
    ws["A4"] = "Ligne P&L"
    ws["B4"] = "Montant (€)"
    for col in ["A4", "B4"]:
        cell = ws[col]
        cell.fill = STYLE_HEADER["fill"]
        cell.font = STYLE_HEADER["font"]
        cell.alignment = Alignment(horizontal="center" if col == "B4" else "left")

    # Largeur colonnes
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 18

    # Structure pour identifier type de chaque ligne
    label_to_type = {label: t for _, label, t in PL_STRUCTURE}
    # Déduplique
    seen = set()
    label_type_dedup = {}
    for _, label, t in PL_STRUCTURE:
        if label not in seen:
            label_type_dedup[label] = t
            seen.add(label)

    # Écriture des lignes
    row = 5
    detail_count = 0

    for label in pl_df.index:
        row_type = label_type_dedup.get(label, "detail")
        montant = pl_df.loc[label, "Montant"] if label in pl_df.index else 0

        cell_a = ws.cell(row=row, column=1, value=label)
        cell_b = ws.cell(row=row, column=2, value=round(montant) if pd.notna(montant) else 0)
        cell_b.number_format = NUMBER_FORMAT
        cell_b.alignment = Alignment(horizontal="right")

        if label in KEY_MARGINS:
            style = STYLE_MARGIN
            cell_a.alignment = Alignment(horizontal="left", indent=0)
            cell_a.border = BORDER_TOTAL
            cell_b.border = BORDER_TOTAL
        elif row_type in ("total", "margin"):
            style = STYLE_TOTAL
            cell_a.alignment = Alignment(horizontal="left", indent=0)
        else:
            # Alternance detail
            style = STYLE_DETAIL if detail_count % 2 == 0 else STYLE_DETAIL_ALT
            cell_a.alignment = Alignment(horizontal="left", indent=2)
            detail_count += 1

        cell_a.fill = style["fill"]
        cell_a.font = style["font"]
        cell_b.fill = style["fill"]
        cell_b.font = style["font"]

        row += 1

    # Freeze panes (fige la ligne d'en-tête)
    ws.freeze_panes = "A5"


# ─────────────────────────────────────────────
# EXPORT EXCEL FINAL
# ─────────────────────────────────────────────

def export_to_excel(pl_dict: dict[str, pd.DataFrame], period: str, output_dir: str = "output"):
    """
    Exporte les 3 P&L dans un fichier Excel multi-onglets.
    
    Args:
        pl_dict: dict {"P&L PID": df, "P&L Celsius": df, "P&L Consolidé": df}
        period: format YYYYMM (ex: "202512")
        output_dir: dossier de sortie
    """
    os.makedirs(output_dir, exist_ok=True)

    # Label période lisible
    p = pd.to_datetime(period, format="%Y%m")
    period_label = p.strftime("%B %Y").capitalize()

    filename = os.path.join(output_dir, f"PL_{period}.xlsx")

    with pd.ExcelWriter(filename, engine="openpyxl") as writer:
        for sheet_name, pl_df in pl_dict.items():
            # Crée l'onglet
            pl_df.to_excel(writer, sheet_name=sheet_name, startrow=3, index=True)
            ws = writer.sheets[sheet_name]

            # Supprime ce qu'a écrit pandas (on réécrit proprement)
            for row in ws.iter_rows():
                for cell in row:
                    cell.value = None

            write_pl_sheet(ws, pl_df, sheet_name, period_label)

    print(f"✅ Export réussi : {filename}")
    return filename